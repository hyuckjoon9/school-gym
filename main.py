"""
근무시간 병합기 - 헬스장 근로자 근무시간 관리 도구
"""

import sys
import os
import re
import pandas as pd
from datetime import datetime

# 구글 시트 연동 관련 import 추가
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials

from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QFormLayout,
    QFileDialog,
    QLabel,
    QMessageBox,
    QLineEdit,
    QComboBox,
    QStackedWidget,
    QSpinBox,
    QGroupBox,
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QPixmap
from PyQt5.QtGui import QIcon
import matplotlib

matplotlib.use("Qt5Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

# 그래프는 한글 깨짐 방지를 위해 Malgun Gothic 사용
plt.rcParams["font.family"] = "Malgun Gothic"
plt.rcParams["axes.unicode_minus"] = False

# exe 배포 시 이미지 파일 경로 처리
if getattr(sys, "frozen", False):
    basedir = sys._MEIPASS
else:
    basedir = os.path.dirname(__file__)

# file URI 접두어를 붙여 절대경로로 처리
# summer_img_path = os.path.join(basedir, "summer.png").replace("\\", "/")
arctic_img_path = os.path.join(basedir, "assets", "img", "gu.png")
icon_path = os.path.join(basedir, "assets", "ico", "zzangu.ico")
env_path = os.path.join(basedir, "config", ".env")


def init_app(self):
    """애플리케이션 초기화 및 심플 화이트 스타일 UI 적용"""
    self.setWindowTitle("근무시간 병합기")
    self.setGeometry(100, 100, 1100, 850)
    self.setWindowIcon(QIcon(icon_path))
    self.setAcceptDrops(True)
    self.setStyleSheet(
        """
        QWidget {
            background-color: #edf5f7;
            font-family: Consolas, monospace;
        }
        QGroupBox {
            background: transparent;
        }
        QPushButton {
            background-color: #FFEDB3;
            border: 2px solid #F4A460;
            border-radius: 8px;
            padding: 8px 16px;
            font-family: Consolas, monospace;
        }
        QPushButton:hover {
            background-color: #FFE066;
        }
        QLineEdit, QComboBox, QSpinBox {
            background-color: #FFFFFF;
            border: 1px solid #CCCCCC;
            border-radius: 4px;
            padding: 4px;
            font-family: Consolas, monospace;
        }
        """
    )
    # 배경 레이블 제거
    if hasattr(self, "bg_label"):
        self.bg_label.hide()
        del self.bg_label


def resizeEvent(self, event):
    """윈도우 크기 변경시 배경 이미지도 같이 리사이즈"""
    super().resizeEvent(event)
    if hasattr(self, "bg_label"):
        self.bg_label.setGeometry(self.rect())
        self.bg_label.lower()


def create_image_widget(self):
    """이미지 위젯 생성: arctic_fox.png 사용"""
    image_label = QLabel(self)
    pixmap = QPixmap(arctic_img_path)
    image_label.setPixmap(pixmap)
    image_label.setAlignment(Qt.AlignCenter)
    image_label.setStyleSheet("margin-bottom: 20px;")
    image_label.setFixedHeight(180)
    image_label.setScaledContents(True)
    image_label.setMaximumWidth(350)
    return image_label


class ScheduleApp(QWidget):
    """근무시간 병합기 메인 애플리케이션"""

    def __init__(self):
        super().__init__()
        self.gcsv_path = None
        self.init_app()
        self.setup_pages()
        self.init_state()

    def init_app(self):
        """애플리케이션 초기화 및 심플 화이트 스타일 UI 적용"""
        self.setWindowTitle("딸깍 딸깍")
        self.setGeometry(100, 100, 1100, 850)
        self.setAcceptDrops(True)
        self.setStyleSheet(
            """
            QWidget {
                background-color: #FFFFFF;
                font-family: Consolas, monospace;
            }
            QGroupBox {
                background: transparent;
            }
            QPushButton {
                background-color: #FFEDB3;
                border: 2px solid #F4A460;
                border-radius: 8px;
                padding: 8px 16px;
                font-family: Consolas, monospace;
            }
            QPushButton:hover {
                background-color: #FFE066;
            }
            QLineEdit, QComboBox, QSpinBox {
                background-color: #FFFFFF;
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 4px;
                font-family: Consolas, monospace;
            }
            """
        )
        # 배경 레이블 제거
        if hasattr(self, "bg_label"):
            self.bg_label.hide()
            del self.bg_label

    def resizeEvent(self, event):
        """윈도우 크기 변경시 배경 이미지도 같이 리사이즈"""
        super().resizeEvent(event)
        if hasattr(self, "bg_label"):
            self.bg_label.setGeometry(self.rect())

    def setup_pages(self):
        """페이지 구성"""
        self.stacked = QStackedWidget(self)
        self.page1 = QWidget()
        self.page2 = QWidget()

        self.init_page1()
        self.init_page2()

        self.stacked.addWidget(self.page1)
        self.stacked.addWidget(self.page2)

        layout = QVBoxLayout(self)
        layout.addWidget(self.stacked)
        self.setLayout(layout)

    def init_state(self):
        """상태 변수 초기화"""
        self.last_stats = None
        self.excel_saved = False
        self.total_hours = 0

    def init_page1(self):
        """페이지 1: 입력 화면"""
        layout = QVBoxLayout()

        # 상단 여백
        layout.addSpacing(20)

        # 이미지 중앙 배치 (상단 중앙)
        image_label = QLabel(self)
        pixmap = QPixmap(arctic_img_path)
        image_label.setPixmap(pixmap)
        image_label.setAlignment(Qt.AlignCenter)
        image_label.setStyleSheet("margin-bottom: 20px;")
        image_label.setFixedHeight(180)
        image_label.setScaledContents(True)
        image_label.setMaximumWidth(350)

        image_row = QHBoxLayout()
        image_row.addStretch()
        image_row.addWidget(image_label)
        image_row.addStretch()
        layout.addLayout(image_row)

        # 여백
        layout.addSpacing(30)

        # 입력 폼
        layout.addWidget(self.create_input_group())
        layout.addSpacing(20)
        layout.addWidget(self.create_file_group())

        # 여백
        layout.addSpacing(40)

        # 버튼
        layout.addLayout(self.create_page1_buttons())

        # 하단 여백
        layout.addSpacing(50)
        layout.addStretch()

        self.page1.setLayout(layout)

    def init_page2(self):
        """페이지 2: 통계 화면"""
        layout = QVBoxLayout()

        # 상단 여백
        layout.addSpacing(20)

        # 통계 및 그래프 (더 큰 공간 할당)
        stats_group = self.create_stats_group()
        layout.addWidget(stats_group, stretch=3)  # 3/4 공간 할당

        layout.addSpacing(15)

        # 시급/월급 (작은 공간 할당)
        wage_group = self.create_wage_group()
        layout.addWidget(wage_group, stretch=1)  # 1/4 공간 할당

        layout.addSpacing(15)

        # 버튼
        layout.addLayout(self.create_page2_buttons())

        # 하단 여백
        layout.addSpacing(20)

        self.page2.setLayout(layout)

    def create_image_widget(self):
        """이미지 위젯 생성"""
        image_label = QLabel(self)
        pixmap = QPixmap(arctic_img_path)
        image_label.setPixmap(pixmap)
        image_label.setAlignment(Qt.AlignCenter)
        image_label.setStyleSheet("margin-bottom: 20px;")
        image_label.setFixedHeight(180)
        image_label.setScaledContents(True)
        image_label.setMaximumWidth(350)
        return image_label

    def create_input_group(self):
        """입력 그룹 생성"""
        group = QGroupBox("근무자 정보 입력")
        group.setStyleSheet("font-size: 20px; font-weight: bold;")
        layout = QFormLayout()

        # 이름 입력
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("예: 권혁준")
        self.name_input.setStyleSheet("font-size: 20px; padding: 6px;")
        self.name_input.setFixedHeight(50)

        # 월 선택
        self.month_combo = QComboBox()
        self.month_combo.addItems(
            [
                "12-1",
                "1-2",
                "2-3",
                "3-4",
                "4-5",
                "5-6",
                "6-7",
                "7-8",
                "8-9",
                "9-10",
                "10-11",
                "11-12",
            ]
        )
        self.month_combo.setStyleSheet("font-size: 20px; padding: 6px;")
        self.month_combo.setFixedHeight(50)

        layout.addRow("이름", self.name_input)
        layout.addRow("근무 월", self.month_combo)
        group.setLayout(layout)
        return group

    def create_file_group(self):
        """파일 선택 그룹 생성 + 구글 시트 버튼 추가"""
        group = QGroupBox("CSV 파일 선택")
        group.setStyleSheet("font-size: 20px; font-weight: bold;")
        layout = QHBoxLayout()

        self.file_path = QLineEdit()
        self.file_path.setReadOnly(True)
        self.file_path.setPlaceholderText("드래그하거나 '찾아보기' 버튼 클릭")
        self.file_path.setStyleSheet("font-size: 20px; padding: 6px;")
        self.file_path.setFixedHeight(50)

        file_button = QPushButton("찾아보기")
        file_button.setFixedWidth(120)
        file_button.setFixedHeight(50)
        file_button.setStyleSheet("font-size: 20px; padding: 8px; font-weight: bold")
        file_button.clicked.connect(self.select_file)

        # 구글 시트에서 가져오기 버튼 추가
        gsheet_button = QPushButton("개쩌는 딸깍")
        gsheet_button.setFixedWidth(180)
        gsheet_button.setFixedHeight(50)
        gsheet_button.setStyleSheet(
            "font-size: 20px; padding: 8px; font-weight: bold; background-color: #c2f0fc;"
        )
        gsheet_button.clicked.connect(self.on_gsheet_btn_clicked)

        layout.addWidget(self.file_path)
        layout.addWidget(file_button)
        layout.addWidget(gsheet_button)
        group.setLayout(layout)
        return group

    def on_gsheet_btn_clicked(self):
        """구글 시트에서 CSV 가져오기"""
        name = self.name_input.text().strip()
        month_range = self.month_combo.currentText().strip()
        if not name or not month_range:
            QMessageBox.warning(self, "오류", "이름과 근무 월을 먼저 입력하세요.")
            return

        try:
            self.download_gsheet_csv(month_range)
            csv_path = self.gcsv_path
            if csv_path and os.path.exists(csv_path):
                self.file_path.setText(csv_path)
                QMessageBox.information(
                    self,
                    "가져오기 완료",
                    f"구글 시트에서 CSV를 가져왔습니다:\n{csv_path}",
                )
            else:
                QMessageBox.warning(self, "오류", "CSV 파일을 가져오지 못했습니다.")
        except Exception as e:
            QMessageBox.critical(self, "구글 시트 오류", str(e))

    def download_gsheet_csv(self, worksheet_name):
        """구글 시트에서 워크시트 데이터를 CSV로 저장하고 경로 반환"""
        # .env 파일 로딩
        # load_dotenv(dotenv_path="./config/.env")
        load_dotenv(dotenv_path=env_path)
        service_account_path = os.getenv("GSHEET_SERVICE_ACCOUNT")
        json_path = os.path.join(basedir, "config", service_account_path)

        SHEET_URL = os.getenv("GSHEET_URL")
        SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

        if not json_path or not SHEET_URL:
            raise Exception(
                "json_path(구글 키 파일) 또는 GSHEET_URL 환경변수 누락/경로 오류"
            )

        creds = Credentials.from_service_account_file(json_path, scopes=SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_url(SHEET_URL)
        worksheet = sh.worksheet(worksheet_name)
        data = worksheet.get_all_values()

        df = pd.DataFrame(data[1:], columns=data[0])
        csv_filename = f"{worksheet_name}_from_api.csv"
        df.to_csv(csv_filename, index=False, encoding="utf-8-sig")
        self.gcsv_path = csv_filename

    def init_page2(self):
        """페이지 2: 통계 화면"""
        layout = QVBoxLayout()

        # 상단 여백
        layout.addSpacing(20)

        # 통계 및 그래프 (더 큰 공간 할당)
        stats_group = self.create_stats_group()
        layout.addWidget(stats_group, stretch=3)  # 3/4 공간 할당

        layout.addSpacing(15)

        # 시급/월급 (작은 공간 할당)
        wage_group = self.create_wage_group()
        layout.addWidget(wage_group, stretch=1)  # 1/4 공간 할당

        layout.addSpacing(15)

        # 버튼
        layout.addLayout(self.create_page2_buttons())

        # 하단 여백
        layout.addSpacing(20)

        self.page2.setLayout(layout)

    def create_stats_group(self):
        """통계 그룹 생성"""
        group = QGroupBox("근무 통계 및 시각화")
        group.setStyleSheet("font-size: 15px; font-weight: bold;")
        layout = QVBoxLayout()

        # 통계 라벨
        self.stats_label = QLabel()
        self.stats_label.setStyleSheet(
            "font-size: 20px; font-weight: bold; margin: 15px;"
        )
        self.stats_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.stats_label)

        # 그래프 (비율 1/4 축소: figsize, 높이, 여백 모두 0.75배)
        self.figure = plt.Figure(figsize=(9, 4.5), dpi=100)  # 12,6 -> 9,4.5
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setMinimumHeight(390)  # 520 -> 390
        layout.addWidget(self.canvas)

        group.setLayout(layout)
        return group

    def create_wage_group(self):
        """시급/월급 그룹 생성"""
        group = QGroupBox("시급 및 월급 계산")
        group.setStyleSheet("font-size: 15px; font-weight: bold;")
        layout = QVBoxLayout()

        # 최저시급 안내
        min_wage_label = QLabel("2025년 최저 시급은 10,030원입니다.")
        min_wage_label.setStyleSheet(
            "font-size: 16px; color: #e67e22; font-weight: bold; margin: 8px;"
        )
        min_wage_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(min_wage_label)

        # 시급 입력
        form_layout = QFormLayout()
        self.wage_input = QSpinBox()
        self.wage_input.setRange(0, 100000)
        self.wage_input.setSingleStep(500)
        self.wage_input.setSuffix(" 원")
        self.wage_input.setValue(10030)
        self.wage_input.setStyleSheet("font-size: 16px; padding: 8px;")
        self.wage_input.valueChanged.connect(self.update_salary)
        form_layout.addRow("시급 입력:", self.wage_input)

        # 월급 표시
        self.salary_label = QLabel("월급: - 원")
        self.salary_label.setStyleSheet(
            "font-size: 18px; color: #1a73e8; font-weight: bold;"
        )
        form_layout.addRow("", self.salary_label)

        layout.addLayout(form_layout)
        group.setLayout(layout)
        return group

    def create_page1_buttons(self):
        """페이지 1 버튼 생성"""
        layout = QHBoxLayout()

        self.run_btn = QPushButton("엑셀로 저장 및 통계 보기")
        self.run_btn.setStyleSheet(
            """
            QPushButton {
                font-size: 18px; 
                padding: 12px 20px; 
                font-weight: bold;
                min-width: 200px;
                max-width: 300px;
            }
        """
        )
        self.run_btn.clicked.connect(self.on_run_btn_clicked)

        layout.addStretch()
        layout.addWidget(self.run_btn)
        layout.addStretch()

        return layout

    def create_page2_buttons(self):
        """페이지 2 버튼 생성 (입력 화면으로 + 다른 사람 선택 버튼 추가)"""
        layout = QHBoxLayout()

        self.back_btn = QPushButton("입력 화면으로")
        self.back_btn.setStyleSheet(
            """
            QPushButton {
                font-size: 16px; 
                padding: 10px 20px;
                min-width: 120px;
                max-width: 180px;
            }
        """
        )
        self.back_btn.clicked.connect(lambda: self.stacked.setCurrentIndex(0))

        # 새로 추가된 '다른 사람 선택' 버튼: 상태 초기화 후 page1로 이동
        self.change_user_btn = QPushButton("다른 사람 선택")
        self.change_user_btn.setStyleSheet(
            """
            QPushButton {
                font-size: 16px; 
                padding: 10px 20px;
                min-width: 120px;
                max-width: 180px;
            }
        """
        )
        self.change_user_btn.clicked.connect(self.reset_app)

        layout.addWidget(self.back_btn)
        layout.addStretch()
        layout.addWidget(self.change_user_btn)
        layout.addStretch()

        return layout

    def reset_app(self):
        """다른 사람 선택 시 전체 입력 및 상태 초기화 후 페이지1 전환"""
        self.name_input.clear()
        self.file_path.clear()
        self.wage_input.setValue(10030)
        self.salary_label.setText("월급: - 원")
        self.stats_label.setText("")
        self.excel_saved = False
        self.last_stats = None
        self.figure.clear()
        self.canvas.draw()
        self.run_btn.setText("엑셀로 저장 및 통계 보기")
        self.stacked.setCurrentIndex(0)

    def on_run_btn_clicked(self):
        """실행 버튼 클릭 처리"""
        if not self.validate_input():
            return

        try:
            name = self.name_input.text().strip()
            path = self.file_path.text().strip()
            stats = self.get_stats_new(path, name)
            if stats is None:
                QMessageBox.warning(
                    self,
                    "오류",
                    "입력한 이름이 CSV에 없습니다. 이름을 다시 확인하세요.",
                )
                if (
                    hasattr(self, "gcsv_path")
                    and self.gcsv_path
                    and os.path.exists(self.gcsv_path)
                ):
                    os.remove(self.gcsv_path)
                return
            self.last_stats = (stats, name, path)

            if not self.excel_saved:
                self.save_to_excel_new(stats, name, path)
                self.excel_saved = True
                self.run_btn.setText("통계 보기")

            self.show_stats_on_page2(stats)
            self.stacked.setCurrentIndex(1)
            if (
                hasattr(self, "gcsv_path")
                and self.gcsv_path
                and os.path.exists(self.gcsv_path)
            ):
                os.remove(self.gcsv_path)

        except Exception as e:
            if (
                hasattr(self, "gcsv_path")
                and self.gcsv_path
                and os.path.exists(self.gcsv_path)
            ):
                os.remove(self.gcsv_path)
            QMessageBox.critical(self, "오류 발생", str(e))

    def validate_input(self):
        """입력 검증"""
        name = self.name_input.text().strip()
        path = self.file_path.text().strip()

        if not name:
            QMessageBox.warning(self, "오류", "근무자 이름을 입력하세요.")
            return False

        if not path or not path.endswith(".csv"):
            QMessageBox.warning(self, "오류", "유효한 CSV 파일을 선택하세요.")
            return False

        return True

    def get_stats(self, file_path, target_name):
        """CSV 파일에서 통계 추출 (요일 포함)"""
        import re
        from datetime import date

        df = pd.read_csv(file_path)
        df.rename(
            columns={df.columns[0]: "날짜", df.columns[1]: "근무자"}, inplace=True
        )
        df.dropna(subset=["날짜", "근무자"], how="all", inplace=True)
        df["날짜"] = df["날짜"].fillna(method="ffill")

        mask = df.applymap(lambda x: target_name in str(x).strip())
        positions = mask.stack()[mask.stack()].index.tolist()
        if not positions:
            return None
        results = [(df.at[row, "날짜"], time) for row, time in positions]

        date_order = []
        for d, _ in results:
            if d not in date_order:
                date_order.append(d)

        grouped = {}
        for d, t in results:
            grouped.setdefault(d, []).append(t)

        def sort_key(t):
            start = t.split("~")[0].strip()
            return datetime.strptime(start, "%H:%M")

        for d in grouped:
            grouped[d].sort(key=sort_key)

        def merge_time_ranges(time_ranges):
            # 동일한 merge 로직. 이번에도 튜플로 반환.
            def to_range(t):
                start_str, end_str = [
                    s.strip() for s in re.sub(r"\s*~\s*", "-", t).split("-")
                ]
                return (
                    datetime.strptime(start_str, "%H:%M"),
                    datetime.strptime(end_str, "%H:%M"),
                )

            ranges = [to_range(t) for t in time_ranges]
            ranges.sort()
            merged = []
            current_start, current_end = ranges[0]
            for start, end in ranges[1:]:
                if start == current_end:
                    current_end = end
                else:
                    merged.append((current_start, current_end))
                    current_start, current_end = start, end
            merged.append((current_start, current_end))
            return merged

        day_list = []
        hours_list = []
        total_minutes = 0

        for d in date_order:
            merged_ranges = merge_time_ranges(grouped[d])
            day_minutes = sum(
                int((end - start).total_seconds() // 60) for start, end in merged_ranges
            )
            total_minutes += day_minutes
            day_list.append(str(d))
            hours_list.append(round(day_minutes / 60, 2))

        # 날짜별 요일 정보 추가하여 day_labels 생성 (예: "7일 (목)")
        weekday_kor = ["월", "화", "수", "목", "금", "토", "일"]
        selected_range = self.month_combo.currentText()  # e.g., "5-6"
        start_month = int(selected_range.split("-")[0])
        cur_year = datetime.now().year
        cur_month = start_month
        prev_day = 0
        day_labels = []
        for d_str in day_list:
            day_int = int(d_str.replace("일", ""))
            if prev_day and day_int < prev_day:
                cur_month += 1
                if cur_month > 12:
                    cur_month = 1
                    cur_year += 1
            prev_day = day_int
            real_date = date(cur_year, cur_month, day_int)
            weekday = weekday_kor[real_date.weekday()]
            day_labels.append(f"{d_str} ({weekday})")

        days = len(day_list)
        total_hours = total_minutes / 60
        avg_hours = total_hours / days if days else 0

        return {
            "days": days,
            "total_hours": total_hours,
            "avg_hours": avg_hours,
            "day_list": day_list,
            "hours_list": hours_list,
            "day_labels": day_labels,
        }

    def get_stats_new(self, file_path, target_name):
        """CSV 파일에서 통계 추출 - 완전히 새로운 로직"""
        import re
        from datetime import datetime, date

        df = pd.read_csv(file_path, header=None)

        # CSV 내용 전체 출력
        print("=== CSV 파일 내용 ===")
        print(df.to_string())
        print("===================")

        # 1단계: 모든 시간 범위 헤더 찾기
        time_headers = {}  # {행번호: {열번호: 시간범위}}

        for idx, row in df.iterrows():
            # 첫 두 열이 비어있고 시간 형식이 있는 행 찾기
            if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
                time_ranges = {}
                for col_idx in range(2, len(row)):
                    cell_value = str(row.iloc[col_idx]).strip()
                    if re.search(r"\d{1,2}:\d{2}\s*[~-]\s*\d{1,2}:\d{2}", cell_value):
                        time_ranges[col_idx] = cell_value

                if time_ranges:  # 시간 범위가 있으면 헤더로 인정
                    time_headers[idx] = time_ranges
                    print(f"시간 헤더 발견 (행 {idx}): {time_ranges}")

        # 2단계: 각 날짜별로 근무자 데이터 수집
        work_data = []  # [(날짜, 시간범위, 헤더행번호)]
        current_header = None
        current_header_row = -1
        current_date = None

        for idx, row in df.iterrows():
            # 새로운 시간 헤더인가?
            if idx in time_headers:
                current_header = time_headers[idx]
                current_header_row = idx
                continue

            # 현재 시간 헤더가 없으면 스킵
            if current_header is None:
                continue

            # 날짜가 있는 행인가? (더 정확한 감지)
            date_cell = str(row.iloc[0]).strip()
            if (
                not pd.isna(row.iloc[0])
                and date_cell
                and date_cell != "nan"
                and "일" in date_cell
            ):
                current_date = date_cell
                print(f"날짜 업데이트: {current_date}")

            # 근무자 행 처리 (current_date가 있을 때만)
            if not pd.isna(row.iloc[1]) and current_date:
                # 해당 근무자의 데이터 찾기
                for col_idx, cell_value in row.items():
                    if isinstance(cell_value, str) and target_name in cell_value:
                        if col_idx in current_header:
                            time_range = current_header[col_idx]
                            work_data.append(
                                (current_date, time_range, current_header_row)
                            )
                            print(
                                f"근무 데이터: {current_date}, {time_range}, 헤더{current_header_row}"
                            )

        if not work_data:
            return None

        # 3단계: 날짜별로 그룹화 및 시간 계산
        def parse_time_range(time_str):
            """시간 범위 문자열을 시작/끝 시간으로 파싱"""
            parts = re.split(r"[~-]", time_str)
            start = datetime.strptime(parts[0].strip(), "%H:%M")
            end = datetime.strptime(parts[1].strip(), "%H:%M")
            return start, end

        def merge_time_ranges(ranges):
            """시간 범위들을 병합"""
            if not ranges:
                return []

            # 시작 시간으로 정렬
            ranges.sort(key=lambda x: x[0])
            merged = [ranges[0]]

            for current_start, current_end in ranges[1:]:
                last_start, last_end = merged[-1]

                # 연속되거나 겹치는 시간이면 병합
                if current_start <= last_end:
                    merged[-1] = (last_start, max(last_end, current_end))
                else:
                    merged.append((current_start, current_end))

            return merged

        # 날짜별로 정리 - CSV 순서대로 유지 (정렬하지 않음)
        date_work = {}  # {날짜: [(시작시간, 끝시간)]}
        date_order = []  # CSV에서 나타난 순서대로

        for work_date, time_range, header_row in work_data:
            if work_date not in date_work:
                date_work[work_date] = []
                date_order.append(work_date)  # 처음 나타난 순서대로 저장

            start_time, end_time = parse_time_range(time_range)
            date_work[work_date].append((start_time, end_time))

        # CSV 순서대로 처리 (정렬하지 않음)
        day_list = []
        hours_list = []
        total_minutes = 0

        for work_date in date_order:  # sorted_dates 대신 date_order 사용
            # 해당 날짜의 시간 범위들을 병합
            merged_ranges = merge_time_ranges(date_work[work_date])

            # 총 근무 시간 계산
            day_minutes = sum(
                (end - start).total_seconds() / 60 for start, end in merged_ranges
            )
            day_hours = day_minutes / 60

            day_list.append(work_date)
            hours_list.append(round(day_hours, 2))
            total_minutes += day_minutes

            print(f"{work_date}: {day_hours:.2f}시간")

        # 요일 정보 추가 - 올바른 월 계산
        weekday_kor = ["월", "화", "수", "목", "금", "토", "일"]
        selected_range = self.month_combo.currentText()  # 예: "8-9"
        start_month = int(selected_range.split("-")[0])  # 8
        next_month = int(selected_range.split("-")[1])  # 9
        cur_year = datetime.now().year

        day_labels = []
        prev_day = 0

        for d_str in day_list:
            day_int = int(re.sub(r"\D", "", d_str))

            # 날짜가 작아지면 다음 월로 넘어간 것
            if prev_day and day_int < prev_day:
                current_month = next_month
            else:
                current_month = start_month

            prev_day = day_int

            # 다음 월이 1월이면 연도 증가
            if current_month == 1 and start_month == 12:
                cur_year += 1

            real_date = date(cur_year, current_month, day_int)
            weekday = weekday_kor[real_date.weekday()]
            day_labels.append(f"{d_str} ({weekday})")
            print(
                f"날짜 처리: {d_str} → {cur_year}년 {current_month}월 {day_int}일 ({weekday})"
            )

        days = len(day_list)
        total_hours = total_minutes / 60
        avg_hours = total_hours / days if days else 0

        print(f"총 근무시간: {total_hours:.2f}시간")

        return {
            "days": days,
            "total_hours": total_hours,
            "avg_hours": avg_hours,
            "day_list": day_list,
            "hours_list": hours_list,
            "day_labels": day_labels,
        }

    def save_to_excel(self, stats, name, csv_path):
        """엑셀 파일 저장 (수정된 로직)"""
        df = pd.read_csv(csv_path)
        df.dropna(subset=["Unnamed: 0", "Unnamed: 1"], how="all", inplace=True)
        df.rename(columns={"Unnamed: 0": "날짜", "Unnamed: 1": "근무자"}, inplace=True)
        df["날짜"] = df["날짜"].fillna(method="ffill")

        mask = df.applymap(lambda x: name in str(x).strip())
        positions = mask.stack()[mask.stack()].index.tolist()
        results = [(df.at[row, "날짜"], time) for row, time in positions]

        date_order = []
        for d, _ in results:
            if d not in date_order:
                date_order.append(d)

        grouped = {}
        for d, t in results:
            grouped.setdefault(d, []).append(t)

        def sort_key(t):
            start = t.split("~")[0].strip()
            return datetime.strptime(start, "%H:%M")

        for d in grouped:
            grouped[d].sort(key=sort_key)

        def merge_time_ranges(time_ranges):
            # 동일한 merge 로직. 이번에도 튜플로 반환.
            def to_range(t):
                start_str, end_str = [
                    s.strip() for s in re.sub(r"\s*~\s*", "-", t).split("-")
                ]
                return (
                    datetime.strptime(start_str, "%H:%M"),
                    datetime.strptime(end_str, "%H:%M"),
                )

            ranges = [to_range(t) for t in time_ranges]
            ranges.sort()
            merged = []
            current_start, current_end = ranges[0]
            for start, end in ranges[1:]:
                if start == current_end:
                    current_end = end
                else:
                    merged.append((current_start, current_end))
                    current_start, current_end = start, end
            merged.append((current_start, current_end))
            return merged

        selected_range = self.month_combo.currentText()  # 예: "5-6"
        start_month = int(selected_range.split("-")[0])  # 앞 숫자만 가져옴
        start_year = datetime.now().year

        weekday_kor = [
            "월요일",
            "화요일",
            "수요일",
            "목요일",
            "금요일",
            "토요일",
            "일요일",
        ]
        cur_year, cur_month, prev_day = start_year, start_month, 0

        records = []
        for d_str in date_order:
            day_int = int(d_str.replace("일", ""))
            if prev_day and day_int < prev_day:
                cur_month += 1
                if cur_month > 12:
                    cur_month = 1
                    cur_year += 1
            prev_day = day_int

            from datetime import date

            real_date = date(cur_year, cur_month, day_int)
            weekday = weekday_kor[real_date.weekday()]
            merged_ranges = merge_time_ranges(grouped[d_str])
            # 문자열로 변환
            merged_times_str = ",".join(
                f"{s.strftime('%H:%M')}-{e.strftime('%H:%M')}" for s, e in merged_ranges
            )

            records.append(
                {
                    "월": cur_month,
                    "일": day_int,
                    "요일": weekday,
                    "근무시간": merged_times_str,
                }
            )

        output_df = pd.DataFrame(records, columns=["월", "일", "요일", "근무시간"])
        file_name = f"{name}.xlsx"
        output_df.to_excel(file_name, index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(file_name)
        ws = wb.active

        font = Font(name="굴림", size=11)
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.font = font

        wb.save(file_name)
        QMessageBox.information(
            self, "저장 완료", f"엑셀 파일이 저장되었습니다:\n{file_name}"
        )

    def save_to_excel_new(self, stats, name, csv_path):
        """엑셀 파일 저장 - 새로운 로직"""
        import re
        from datetime import datetime, date

        df = pd.read_csv(csv_path, header=None)

        # 1단계: 모든 시간 범위 헤더 찾기
        time_headers = {}  # {행번호: {열번호: 시간범위}}

        for idx, row in df.iterrows():
            # 첫 두 열이 비어있고 시간 형식이 있는 행 찾기
            if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
                time_ranges = {}
                for col_idx in range(2, len(row)):
                    cell_value = str(row.iloc[col_idx]).strip()
                    if re.search(r"\d{1,2}:\d{2}\s*[~-]\s*\d{1,2}:\d{2}", cell_value):
                        time_ranges[col_idx] = cell_value

                if time_ranges:  # 시간 범위가 있으면 헤더로 인정
                    time_headers[idx] = time_ranges

        # 2단계: 각 날짜별로 근무자 데이터 수집
        work_data = []  # [(날짜, 시간범위, 헤더행번호)]
        current_header = None
        current_header_row = -1
        current_date = None

        for idx, row in df.iterrows():
            # 새로운 시간 헤더인가?
            if idx in time_headers:
                current_header = time_headers[idx]
                current_header_row = idx
                continue

            # 현재 시간 헤더가 없으면 스킵
            if current_header is None:
                continue

            # 날짜가 있는 행인가? (더 정확한 감지)
            date_cell = str(row.iloc[0]).strip()
            if (
                not pd.isna(row.iloc[0])
                and date_cell
                and date_cell != "nan"
                and "일" in date_cell
            ):
                current_date = date_cell

            # 근무자 행 처리 (current_date가 있을 때만)
            if not pd.isna(row.iloc[1]) and current_date:
                # 해당 근무자의 데이터 찾기
                for col_idx, cell_value in row.items():
                    if isinstance(cell_value, str) and name in cell_value:
                        if col_idx in current_header:
                            time_range = current_header[col_idx]
                            work_data.append(
                                (current_date, time_range, current_header_row)
                            )

        if not work_data:
            QMessageBox.warning(
                self, "오류", "해당 근무자의 데이터를 찾을 수 없습니다."
            )
            return

        # 3단계: 날짜별로 정리하고 엑셀 형식으로 변환
        def parse_time_range(time_str):
            """시간 범위 문자열을 시작/끝 시간으로 파싱"""
            parts = re.split(r"[~-]", time_str)
            start = datetime.strptime(parts[0].strip(), "%H:%M")
            end = datetime.strptime(parts[1].strip(), "%H:%M")
            return start, end

        def merge_time_ranges(ranges):
            """시간 범위들을 병합"""
            if not ranges:
                return []

            # 시작 시간으로 정렬
            ranges.sort(key=lambda x: x[0])
            merged = [ranges[0]]

            for current_start, current_end in ranges[1:]:
                last_start, last_end = merged[-1]

                # 연속되거나 겹치는 시간이면 병합
                if current_start <= last_end:
                    merged[-1] = (last_start, max(last_end, current_end))
                else:
                    merged.append((current_start, current_end))

            return merged

        # 날짜별로 정리 - CSV 순서대로 유지
        date_work = {}  # {날짜: [(시작시간, 끝시간)]}
        date_order = []  # CSV에서 나타난 순서대로

        for work_date, time_range, header_row in work_data:
            if work_date not in date_work:
                date_work[work_date] = []
                date_order.append(work_date)  # 처음 나타난 순서대로 저장

            start_time, end_time = parse_time_range(time_range)
            date_work[work_date].append((start_time, end_time))

        selected_range = self.month_combo.currentText()  # 예: "8-9"
        start_month = int(selected_range.split("-")[0])  # 8
        next_month = int(selected_range.split("-")[1])  # 9
        start_year = datetime.now().year

        weekday_kor = [
            "월요일",
            "화요일",
            "수요일",
            "목요일",
            "금요일",
            "토요일",
            "일요일",
        ]

        records = []
        prev_day = 0
        is_next_month = False  # 다음 월로 넘어갔는지 추적

        for work_date in date_order:  # sorted_dates 대신 date_order 사용
            day_int = int(re.sub(r"\D", "", work_date))

            # 날짜가 작아지면 다음 월로 넘어간 것
            if prev_day and day_int < prev_day:
                is_next_month = True

            # 월과 연도 결정
            if is_next_month:
                current_month = next_month
                current_year = start_year
                # 다음 월이 1월이면 연도 증가
                if current_month == 1 and start_month == 12:
                    current_year += 1
            else:
                current_month = start_month
                current_year = start_year

            prev_day = day_int

            real_date = date(current_year, current_month, day_int)
            weekday = weekday_kor[real_date.weekday()]

            # 해당 날짜의 시간 범위들을 병합
            merged_ranges = merge_time_ranges(date_work[work_date])

            # 시간 문자열 생성
            time_strs = []
            for start_time, end_time in merged_ranges:
                time_str = (
                    f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"
                )
                time_strs.append(time_str)

            merged_times_str = ",".join(time_strs)

            records.append(
                {
                    "월": current_month,
                    "일": day_int,
                    "요일": weekday,
                    "근무시간": merged_times_str,
                }
            )

            print(f"엑셀: {work_date} → {current_year}년 {current_month}월 {day_int}일")

        output_df = pd.DataFrame(records, columns=["월", "일", "요일", "근무시간"])
        file_name = f"{name}_{selected_range}.xlsx"
        output_df.to_excel(file_name, index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(file_name)
        ws = wb.active

        font = Font(name="굴림", size=11)
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.font = font

        wb.save(file_name)
        QMessageBox.information(
            self, "저장 완료", f"엑셀 파일이 저장되었습니다:\n{file_name}"
        )

    def update_salary(self):
        """월급 업데이트"""
        wage = self.wage_input.value()
        if hasattr(self, "total_hours"):
            salary = int(self.total_hours * wage)
            self.salary_label.setText(f"월급: <b>{salary:,} 원</b>")
        else:
            self.salary_label.setText("월급: - 원")

    def select_file(self):
        """파일 선택 다이얼로그"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "CSV 파일 선택", "", "CSV Files (*.csv)"
        )
        if file_path:
            self.file_path.setText(file_path)

    def dragEnterEvent(self, event):
        """드래그 이벤트 처리"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        """드롭 이벤트 처리"""
        urls = event.mimeData().urls()
        if urls:
            file_path = urls[0].toLocalFile()
            if file_path.endswith(".csv"):
                self.file_path.setText(file_path)

    def show_stats_on_page2(self, stats):
        """페이지 2에 통계와 그래프 업데이트"""
        self.stats_label.setText(
            f"<span style='font-size:20px;'>"
            f"총 근무일수: <b>{stats['days']}</b>일, "
            f"총 근무시간: <b>{stats['total_hours']:.2f}</b>시간, "
            f"평균 1일 근무시간: <b>{stats['avg_hours']:.2f}</b>시간"
            f"</span>"
        )
        self.total_hours = stats["total_hours"]
        self.update_salary()
        self.create_chart(stats)

    def create_chart(self, stats):
        """차트 생성 (정보가 잘리지 않도록 내부 여백 충분히 확보)"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)

        day_labels = stats.get("day_labels", stats["day_list"])
        hours = stats["hours_list"]

        bars = ax.bar(range(len(day_labels)), hours, color="#4a90e2", width=0.5)

        ax.set_title(
            "일별 근무시간",
            fontdict={"fontsize": 14, "fontfamily": "Malgun Gothic"},
            pad=26,
        )
        ax.set_xlabel(
            "날짜",
            fontdict={"fontsize": 12, "fontfamily": "Malgun Gothic"},
            labelpad=7,
        )
        ax.set_ylabel(
            "근무시간(시간)",
            fontdict={"fontsize": 12, "fontfamily": "Malgun Gothic"},
            labelpad=7,
        )

        ax.set_xticks(range(len(day_labels)))
        ax.set_xticklabels(
            day_labels, rotation=45, fontsize=9, fontfamily="Malgun Gothic"
        )
        ax.tick_params(axis="y", labelsize=9)

        max_hours = max(hours) if hours else 0
        ax.set_ylim(0, max_hours * 1.1)

        for bar, h in zip(bars, hours):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                h + max(hours) * 0.01,
                f"{h:.1f}h",
                ha="center",
                va="bottom",
                fontsize=8,
                fontfamily="Malgun Gothic",
                color="#333",
                weight="bold",
            )

        # 내부 플롯 영역을 줄여 상하좌우 여백을 넉넉히 확보 (정보가 안 짤리게!)
        self.figure.tight_layout(pad=2.25)
        self.figure.subplots_adjust(
            left=0.13,  # 좌측 여백 약간 늘림
            right=0.97,  # 우측 여백 약간 늘림
            bottom=0.28,  # 하단 여백 충분히 확보 (x축 날짜 안 짤리게)
            top=0.80,  # 상단 여백 충분히 확보 (제목 안 짤리게)
        )

        self.canvas.draw()


def main():
    """메인 함수"""
    app = QApplication(sys.argv)
    window = ScheduleApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
